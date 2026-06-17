"""Build the H-1B sponsor index from US DoL OFLC LCA disclosure files.

The DoL publishes quarterly LCA disclosure workbooks (free, official) at
https://www.dol.gov/agencies/eta/foreign-labor/performance — one big .xlsx per quarter, ~1M
rows, 75+ columns. This is a *build-time* ETL: download one or more of those workbooks, then::

    python scripts/build_h1b_sponsors.py ~/Downloads/LCA_Disclosure_Data_FY2025_Q*.xlsx

It streams the rows (openpyxl read-only, so the giant file never loads fully into memory), keeps
employers whose CASE_STATUS is *certified*, normalizes the employer name the same way dedup does
(so "STRIPE, INC." -> "stripe" matches a "Stripe" posting), counts certified filings, and writes
a compact ``src/ergon_tracker/registry/data/h1b_sponsors.json`` consumed by ``extract/visa.py``.

The parsing logic (``sponsors_from_rows``) is a pure function over row dicts, so it is unit-tested
without needing a real workbook.
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from collections.abc import Iterable, Iterator
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from ergon_tracker.dedup import normalize_company  # noqa: E402

OUT = ROOT / "src" / "ergon_tracker" / "registry" / "data" / "h1b_sponsors.json"

# Header names vary slightly across fiscal years; match case-insensitively against these.
_EMPLOYER_COLS = ("EMPLOYER_NAME", "EMPLOYER_BUSINESS_DBA", "EMPLOYER_LEGAL_BUSINESS_NAME")
_STATUS_COLS = ("CASE_STATUS",)
# Only certified LCAs count as demonstrated sponsorship ("Certified", "Certified - Withdrawn").
_CERTIFIED_PREFIX = "certified"


def _pick(row: dict[str, object], candidates: tuple[str, ...]) -> object | None:
    lowered = {k.lower(): v for k, v in row.items() if isinstance(k, str)}
    for cand in candidates:
        if cand.lower() in lowered:
            return lowered[cand.lower()]
    return None


def sponsors_from_rows(rows: Iterable[dict[str, object]]) -> dict[str, int]:
    """Pure: rows (header->value dicts) -> {normalized_employer_name: certified_filing_count}."""
    counts: Counter[str] = Counter()
    for row in rows:
        status = _pick(row, _STATUS_COLS)
        if not isinstance(status, str) or not status.strip().lower().startswith(_CERTIFIED_PREFIX):
            continue
        employer = _pick(row, _EMPLOYER_COLS)
        if not isinstance(employer, str) or not employer.strip():
            continue
        key = normalize_company(employer)
        if key:
            counts[key] += 1
    return dict(counts)


def read_xlsx_rows(path: Path) -> Iterator[dict[str, object]]:
    """Stream rows of an LCA workbook as header->value dicts (read-only, memory-bounded)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows)]
        for values in rows:
            yield dict(zip(header, values, strict=False))
    finally:
        wb.close()


def main(paths: list[str]) -> None:
    if not paths:
        print(__doc__)
        print("\nUsage: python scripts/build_h1b_sponsors.py <LCA_Disclosure_*.xlsx> [more ...]")
        raise SystemExit(2)

    total: Counter[str] = Counter()
    for p in paths:
        path = Path(p).expanduser()
        if not path.exists():
            print(f"  ! missing: {path}")
            continue
        print(f"  reading {path.name} ...")
        part = sponsors_from_rows(read_xlsx_rows(path))
        for name, n in part.items():
            total[name] += n
        print(f"    +{len(part):,} employers (running total {len(total):,})")

    payload = {
        "source": "US DoL OFLC LCA disclosure data (certified filings)",
        "count": len(total),
        # name -> certified filing count; membership is what extract/visa.py needs.
        "sponsors": dict(sorted(total.items())),
    }
    OUT.write_text(json.dumps(payload), encoding="utf-8")
    print(f"\nwrote {OUT.relative_to(ROOT)} — {len(total):,} unique sponsors")


if __name__ == "__main__":
    main(sys.argv[1:])
